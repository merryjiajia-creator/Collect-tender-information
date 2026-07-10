#!/usr/bin/env python3
"""将 store.json 同步到 4 个腾讯在线表格（追加模式，不覆盖已有数据）。

流程：读取在线表格现有数据 → 与 store.json 合并去重 → 生成 xlsx → import 覆盖。
虽然每次是全量 import，但数据合并确保了历史数据不丢失。
"""
import json, subprocess, hashlib, os, time, sys

TOKEN = os.environ.get("TENCENT_DOCS_TOKEN", "58ad6686deef480e904dd6d323a77b90")
STORE_PATH = os.environ.get("STORE_PATH", "data/store.json")

# 固定4个表格 file_id
TABLES = {
    "总表": "BjidfJJgukTc",
    "物码": "BfTjFAbCdxhP",
    "即时零售": "BjPrfNnnCbOW",
    "到店": "BCcUtVxtSChk",
}

HEADER = ["序号","类别","信息来源平台","招标标题","招标单位","发布时间",
          "截止时间","预算金额","原文链接","项目地点","采购内容（信息摘要）",
          "联系人以及联系方式","跟进团队","跟进人","跟进状态","备注"]

# store.json 字段 → 表格列名
STORE_TO_TABLE = {
    "类别": "类别", "信息来源平台": "信息来源平台", "招标标题": "招标标题",
    "招标单位": "招标单位", "发布时间": "发布时间", "截止时间": "截止时间",
    "预算金额": "预算金额", "资料来源": "原文链接", "项目地点": "项目地点",
    "采购内容": "采购内容（信息摘要）", "联系方式": "联系人以及联系方式",
    "跟进团队": "跟进团队", "跟进人": "跟进人", "跟进状态": "跟进状态",
    "备注": "备注"
}


def call_mcp(name, args):
    payload = {"jsonrpc": "2.0", "id": 1, "method": "tools/call", "params": {"name": name, "arguments": args}}
    r = subprocess.run(["curl", "-s", "-X", "POST", "https://docs.qq.com/openapi/mcp",
        "-H", f"Authorization: {TOKEN}", "-H", "Content-Type: application/json",
        "-d", json.dumps(payload, ensure_ascii=False)], capture_output=True, text=True)
    return json.loads(r.stdout)


def read_table(fid):
    """从在线表格读取所有数据行，返回 list[dict]"""
    result = call_mcp("get_content", {"file_id": fid})
    content = result.get('result', {}).get('structuredContent', {}).get('content', '')
    lines = content.split('\n')
    records = []
    for l in lines:
        if not l.startswith('|') or l.startswith('|--') or '序号' in l or '招标信息' in l:
            continue
        cols = [c.strip() for c in l.split('|')[1:-1]]
        if len(cols) >= 16:
            rec = {HEADER[i]: cols[i] for i in range(min(16, len(cols)))}
            records.append(rec)
    return records


def merge_and_upload(name, fid, store_records):
    """合并 store 数据与表格现有数据，生成 xlsx 并上传"""
    # 读取现有
    existing = read_table(fid)
    print(f"  {name}: 在线表格现有 {len(existing)} 条")

    # 用原文链接去重
    existing_urls = set()
    for r in existing:
        url = r.get('原文链接', '').strip()
        if url:
            existing_urls.add(url)
    
    # store 中属于该类别的记录
    cat_filter = {"总表": None, "物码": "物码", "即时零售": "即时零售", "到店": "到店"}
    cat = cat_filter[name]
    
    new_from_store = []
    for r in store_records:
        if cat and r.get('类别') != cat:
            continue
        url = r.get('资料来源', '').strip()
        if url and url not in existing_urls:
            new_from_store.append(r)
            existing_urls.add(url)
    
    print(f"  {name}: store 新增 {len(new_from_store)} 条")

    # 合并：现有数据 + store 新增
    all_records = existing + new_from_store
    
    # 构建表格行
    rows = []
    for i, rec in enumerate(all_records):
        row = []
        for col in HEADER:
            if col == "序号":
                row.append(str(i + 1))
            elif col == "原文链接":
                # 兼容两个来源
                val = rec.get("原文链接", "") or rec.get("资料来源", "")
                row.append(str(val) if val else "")
            else:
                field = STORE_TO_TABLE.get(col, col)
                val = rec.get(field, "") or rec.get(col, "")
                row.append(str(val) if val is not None else "")
        rows.append(row)
    
    if not rows:
        print(f"  {name}: 无数据，跳过上传")
        return
    
    # 生成 xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    
    wb = Workbook()
    ws = wb.active
    ws.title = "招标信息"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    for col_idx, col_name in enumerate(HEADER, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row_idx, row in enumerate(rows):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx + 2, column=col_idx, value=val)
    
    # 跟进状态下拉
    if rows:
        dv = DataValidation(type="list", formula1='"待跟进,跟进中,述标中,已中标,未中标"', allow_blank=True)
        dv.error = "请选择：待跟进/跟进中/述标中/已中标/未中标"
        ws.add_data_validation(dv)
        dv.add(f"O2:O{len(rows)+1}")
    
    col_widths = [6, 8, 14, 40, 22, 12, 18, 12, 40, 14, 45, 30, 10, 10, 10, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = w
    
    fpath = f"/tmp/sync_{name}.xlsx"
    wb.save(fpath)
    
    # 上传
    fsize = os.path.getsize(fpath)
    with open(fpath, 'rb') as f:
        md5 = hashlib.md5(f.read()).hexdigest()
    
    result = call_mcp("manage.pre_import", {
        "file_name": f"{name}.xlsx", "file_size": fsize, "file_md5": md5
    })
    sc = result.get('result', {}).get('structuredContent', {})
    upload_url = sc.get('upload_url', '')
    file_key = sc.get('file_key', '')
    task_id = sc.get('task_id', '')
    
    if not upload_url:
        print(f"  ❌ pre_import failed: {sc.get('error')}")
        return
    
    subprocess.run(["curl", "-s", "-X", "PUT", "-T", fpath, upload_url], capture_output=True)
    
    call_mcp("manage.async_import", {
        "file_size": fsize, "task_id": task_id, "file_key": file_key,
        "file_name": f"{name}.xlsx", "file_md5": md5
    })
    
    for _ in range(15):
        time.sleep(2)
        result4 = call_mcp("manage.import_progress", {"task_id": task_id})
        sc4 = result4.get('result', {}).get('structuredContent', {})
        if sc4.get('progress') == 100:
            new_fid = sc4.get('file_id', '')
            new_url = sc4.get('file_url', '')
            print(f"  ✅ 上传成功: {new_fid} → {new_url}")
            # 返回新的 file_id（注意：import 会创建新文件！）
            return new_fid
        elif sc4.get('error'):
            print(f"  ❌ 导入失败: {sc4.get('error')}")
            return None
    
    print(f"  ⚠️ 导入超时")
    return None


def main():
    # 加载 store.json
    with open(STORE_PATH) as f:
        store = json.load(f)
    
    # 过滤：只保留符合日期规则的记录
    from datetime import datetime, date
    cutoff = date(2026, 6, 1)
    today = date.today()
    filtered = []
    for r in store:
        pub = None
        pub_str = r.get('发布时间', '')
        for fmt in ['%Y-%m-%d', '%Y/%m/%d']:
            try:
                pub = datetime.strptime(pub_str[:10], fmt).date()
                break
            except: pass
        deadline = None
        dl_str = r.get('截止时间', '')
        for fmt in ['%Y-%m-%d', '%Y/%m/%d']:
            try:
                deadline = datetime.strptime(dl_str[:10], fmt).date()
                break
            except: pass
        if pub and pub < cutoff:
            continue
        if deadline and deadline < today:
            continue
        filtered.append(r)
    
    print(f"store.json: {len(store)} 条 → 过滤后 {len(filtered)} 条")
    
    new_ids = {}
    for name, fid in TABLES.items():
        new_fid = merge_and_upload(name, fid, filtered)
        if new_fid:
            new_ids[name] = new_fid
    
    # 输出新的 file_id 映射
    print("\n=== 更新后的表格 ID ===")
    for name, fid in new_ids.items():
        print(f"  {name}: {fid}")
    
    return new_ids


if __name__ == "__main__":
    main()
